"""Reusable openpyxl template-fill helpers.

Both the price-proposal generator (`proposal_service`) and the production-
instructions generator (`production_service`) fill an .xlsx template that uses
literal placeholder strings as the contract between the file and the code:
one "anchor" row inside an Excel Table is the styling prototype every data
line is copied from, and rows below it (totals / footer / signatures) are
pushed down with `insert_rows`.

openpyxl drops several things on save that these templates rely on — merged
ranges aren't shifted, custom row heights stay keyed at the old rows, and the
print-time header/footer images (logo + footer VML drawings) are discarded.
The helpers here repair all of that so the saved workbook matches the
template's intent.
"""
from __future__ import annotations

import io
import re
import xml.etree.ElementTree as ET
from copy import copy
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl.utils import get_column_letter


# A bare cell reference (col + row) with optional $ anchors. The negative
# lookbehind keeps us from matching inside table/structured names or numbers.
CELL_REF_RE = re.compile(r'(?<![A-Za-z0-9_$])(\$?)([A-Z]+)(\$?)(\d+)')

# SUM(<col>X:<col>Y) with optional $ anchors — used to detect the column-total
# formulas in the original totals row so we can re-aim them at the new block.
SUM_RE = re.compile(r'^=SUM\((\$?)([A-Z]+)(\$?)(\d+):(\$?)([A-Z]+)(\$?)(\d+)\)$', re.IGNORECASE)


def col_to_idx(letter: str) -> int:
    return ord(letter.upper()) - ord('A') + 1


def shift_formula_refs(formula: str, threshold_row: int, shift: int) -> str:
    """Shift the row component of every cell reference whose row is >= threshold."""
    def repl(m):
        col_abs, col, row_abs, row = m.group(1), m.group(2), m.group(3), int(m.group(4))
        if row >= threshold_row:
            return f'{col_abs}{col}{row_abs}{row + shift}'
        return m.group(0)
    return CELL_REF_RE.sub(repl, formula)


def copy_cell_style(src, dst) -> None:
    if not src.has_style:
        return
    dst.font          = copy(src.font)
    dst.fill          = copy(src.fill)
    dst.border        = copy(src.border)
    dst.alignment     = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection    = copy(src.protection)


def extend_print_area(ws, anchor_row: int, n: int) -> None:
    """Push the print area's bottom row down by `n - 1` so it still covers the
    same template content (totals, legal text, signature) after our row
    insertion. Column range is left untouched, per spec."""
    if n <= 1:
        return
    pa = ws.print_area
    if not pa:
        return
    pa_str = pa[0] if isinstance(pa, list) else pa
    if not pa_str:
        return
    # Print area strings come back like "'pricing'!$A$1:$I$28" — peel off the sheet prefix.
    sheet_prefix = ''
    body = pa_str
    if '!' in body:
        sheet_prefix, body = body.rsplit('!', 1)
        sheet_prefix += '!'
    from openpyxl.utils.cell import range_boundaries, get_column_letter
    try:
        min_col, min_row, max_col, max_row = range_boundaries(body.replace('$', ''))
    except Exception:
        return
    if max_row < anchor_row:
        return  # print area sits entirely above the data block; nothing to grow.
    new_max_row = max_row + (n - 1)
    ws.print_area = (
        f"{sheet_prefix}${get_column_letter(min_col)}${min_row}:"
        f"${get_column_letter(max_col)}${new_max_row}"
    )


def extend_excel_table(ws, anchor_row: int, n: int) -> None:
    """Grow any Excel Table whose first data row equals `anchor_row` to span
    `n` data rows. Excel applies the table's style (row stripes) and the
    calculated-column formulas to every row inside `table.ref`, so this is
    what makes new rows inherit the template's table formatting + auto-totals.
    """
    from openpyxl.utils.cell import range_boundaries, get_column_letter
    last_data_row = anchor_row + n - 1
    for table_name in list(ws.tables):
        table = ws.tables[table_name]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        except Exception:
            continue
        first_data = min_row + (table.headerRowCount or 1)
        if first_data != anchor_row:
            continue
        new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_data_row}"
        table.ref = new_ref
        if table.autoFilter is not None:
            table.autoFilter.ref = new_ref


def set_table_range(ws, header_row: int, last_data_row: int) -> None:
    """Resize every Excel Table whose header is at `header_row` so it spans
    down to `last_data_row`. Use when a single table holds several independently
    filled sections (so its first data row is NOT a section anchor and
    `extend_excel_table` can't match it) — call once after all sections fill."""
    from openpyxl.utils.cell import range_boundaries, get_column_letter
    for table_name in list(ws.tables):
        table = ws.tables[table_name]
        try:
            min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
        except Exception:
            continue
        if min_row != header_row:
            continue
        new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_data_row}"
        table.ref = new_ref
        if table.autoFilter is not None:
            table.autoFilter.ref = new_ref


def grow_print_area_to(ws, last_row: int) -> None:
    """Ensure the sheet's print area extends down to at least `last_row`,
    keeping its column range and top row. No-op if already covered."""
    pa = ws.print_area
    if not pa:
        return
    pa_str = pa[0] if isinstance(pa, list) else pa
    if not pa_str:
        return
    sheet_prefix = ''
    body = pa_str
    if '!' in body:
        sheet_prefix, body = body.rsplit('!', 1)
        sheet_prefix += '!'
    from openpyxl.utils.cell import range_boundaries, get_column_letter
    try:
        min_col, min_row, max_col, max_row = range_boundaries(body.replace('$', ''))
    except Exception:
        return
    if max_row >= last_row:
        return
    ws.print_area = (
        f"{sheet_prefix}${get_column_letter(min_col)}${min_row}:"
        f"${get_column_letter(max_col)}${last_row}"
    )


def shift_row_dimensions(ws, insert_row: int, shift: int, anchor_row: int) -> None:
    """openpyxl 3.1's insert_rows leaves `row_dimensions` keyed at the
    original row numbers, which means custom row heights below the insert
    point end up on the wrong rows (e.g. the 74pt legal-paragraph row stays
    at row 28 even though its content moved to row 36). Shift them manually,
    then copy the anchor row's height onto every newly-inserted row so the
    table body is uniform.
    """
    if shift <= 0:
        return
    # Walk highest-first to avoid colliding with rows we're about to overwrite.
    for r in sorted([k for k in list(ws.row_dimensions) if k >= insert_row], reverse=True):
        dim = ws.row_dimensions[r]
        del ws.row_dimensions[r]
        new_r = r + shift
        dim.r = new_r
        ws.row_dimensions[new_r] = dim

    anchor_dim = ws.row_dimensions.get(anchor_row)
    if anchor_dim is not None and anchor_dim.height is not None:
        # `RowDimension.customHeight` is a read-only property that's True iff
        # `height` is set, so just assigning height is enough.
        for r in range(insert_row, insert_row + shift):
            ws.row_dimensions[r].height = anchor_dim.height


def shift_merged_ranges(ws, insert_row: int, shift: int) -> None:
    """Move every merged range that starts at or below `insert_row` down by
    `shift`. openpyxl 3.1's insert_rows leaves merge ranges unshifted, which
    causes writes to cells trapped under a stale merge to silently disappear.
    Call this BEFORE insert_rows so the merges land at the right rows once
    the row insertion completes.
    """
    if shift <= 0:
        return
    to_shift = [mr for mr in ws.merged_cells.ranges if mr.min_row >= insert_row]
    for mr in to_shift:
        c1 = get_column_letter(mr.min_col)
        c2 = get_column_letter(mr.max_col)
        ws.unmerge_cells(str(mr))
        ws.merge_cells(f'{c1}{mr.min_row + shift}:{c2}{mr.max_row + shift}')


def replace_placeholders(ws, placeholder_fields: set[str], ctx: dict) -> None:
    """Replace every static placeholder token (e.g. CUSTOMER_NAME) found in a
    cell with its value from `ctx`."""
    for excel_row in ws.iter_rows():
        for cell in excel_row:
            if cell.value is None:
                continue
            v = str(cell.value).strip()
            if v in placeholder_fields:
                cell.value = ctx.get(v, '')


def find_anchor_row(ws, anchor_placeholder: str, *, clear: bool = True) -> int | None:
    """Return the row that holds `anchor_placeholder`, clearing the cell so the
    first data row can overwrite it. Returns None if not found."""
    for excel_row in ws.iter_rows():
        for cell in excel_row:
            if cell.value is None:
                continue
            if str(cell.value).strip() == anchor_placeholder:
                row = cell.row
                if clear:
                    cell.value = None
                return row
    return None


def fill_section(ws, anchor_row: int, rows: list[dict], column_map: dict, *, extend_table: bool = True) -> int:
    """Insert + fill `rows` starting at `anchor_row` (already located & cleared).

    Replicates the prototype (anchor) row's formatting across every new row,
    pushes the content below the anchor down with `insert_rows`, replicates any
    [#This Row] table formulas the prototype carried, and re-aims SUM / relative
    formulas in the shifted block. Returns the last data row index.

    `extend_table=False` skips growing the owning Excel Table / print area —
    use it when one table holds several sections and the caller resizes the
    table once after all sections are filled.
    """
    n = len(rows)
    if n == 0:
        return anchor_row

    # Per-row formulas the template author placed in the prototype row (e.g.
    # column I = unit price × qty as a Table structured ref). These use
    # [#This Row] references, so the identical text is valid on every data row.
    anchor_formulas = {
        cell.column: cell.value
        for cell in ws[anchor_row]
        if isinstance(cell.value, str) and cell.value.startswith('=')
    }

    shift = n - 1
    last_data_row = anchor_row + n - 1

    captured_formulas: list[tuple[int, int, str]] = []
    if shift > 0:
        # openpyxl doesn't rewrite formula text on insert_rows, so capture every
        # formula below the anchor BEFORE inserting; re-apply them with shifted
        # row references afterwards.
        for r in range(anchor_row + 1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.startswith('='):
                    captured_formulas.append((r, c, v))
        shift_merged_ranges(ws, anchor_row + 1, shift)
        shift_row_dimensions(ws, anchor_row + 1, shift, anchor_row)
        ws.insert_rows(anchor_row + 1, amount=shift)
        # Inserted rows are blank-styled — copy the anchor row's formatting.
        for offset in range(1, n):
            target = anchor_row + offset
            for c in range(1, ws.max_column + 1):
                copy_cell_style(ws.cell(anchor_row, c), ws.cell(target, c))

    # Fill data rows. Formula columns get the prototype formula replicated;
    # empty/None values are left blank.
    for offset, row in enumerate(rows):
        excel_row_idx = anchor_row + offset
        for key, col_letter in column_map.items():
            col_idx = col_to_idx(col_letter)
            if col_idx in anchor_formulas:
                ws.cell(excel_row_idx, col_idx).value = anchor_formulas[col_idx]
                continue
            val = row.get(key)
            ws.cell(excel_row_idx, col_idx).value = None if (val == '' or val is None) else val

    if extend_table:
        extend_excel_table(ws, anchor_row, n)
        extend_print_area(ws, anchor_row, n)

    # Rewrite captured formulas at their new positions: SUM placeholders grow to
    # span the full data block; every other formula keeps shape but shifts rows.
    for orig_r, orig_c, formula in captured_formulas:
        new_r = orig_r + shift
        cell = ws.cell(new_r, orig_c)
        m = SUM_RE.match(formula)
        if m and m.group(2) == m.group(6):
            col = m.group(2)
            cell.value = f'=SUM({col}{anchor_row}:{col}{last_data_row})'
        else:
            cell.value = shift_formula_refs(formula, threshold_row=anchor_row + 1, shift=shift)

    return last_data_row


# ───────────────────────────────────────────────────────────────────────
# Header/footer image restoration (post-openpyxl-save)
# ───────────────────────────────────────────────────────────────────────

# openpyxl reads & re-emits the worksheet's `<headerFooter><oddHeader>&C&G…`
# placeholder text but drops everything else needed to make the print-time
# logo + footer image actually render: the VML drawing files at
# xl/drawings/vmlDrawing*.vml, the image media files, the per-sheet
# relationship to the VML drawing, and the `<legacyDrawingHF r:id="…"/>`
# element inside each worksheet. Without all four, Excel finds nothing for
# `&G` to point at and prints a blank header/footer.
#
# The fix: after openpyxl's serializer runs, copy the missing parts from the
# original template zip into the output zip (matching by sheet name so the
# right VML lands on the right sheet), and patch each output sheet's XML +
# rels to re-link them.

_REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
_MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'


def _parse_workbook_sheets(xml_bytes: bytes) -> list[tuple[str, str]]:
    """Return [(sheet_name, r:id), …] in order from a workbook.xml."""
    root = ET.fromstring(xml_bytes)
    out = []
    for s in root.iter(f'{{{_MAIN_NS}}}sheet'):
        out.append((s.attrib['name'], s.attrib[f'{{{_R_NS}}}id']))
    return out


def _parse_rels(xml_bytes: bytes) -> list[dict]:
    root = ET.fromstring(xml_bytes)
    return [dict(r.attrib) for r in root.iter(f'{{{_REL_NS}}}Relationship')]


def restore_header_footer_images(template_bytes: bytes, output_bytes: bytes) -> bytes:
    template_files: dict[str, bytes] = {}
    with ZipFile(io.BytesIO(template_bytes), 'r') as tz:
        for name in tz.namelist():
            template_files[name] = tz.read(name)
    output_files: dict[str, bytes] = {}
    with ZipFile(io.BytesIO(output_bytes), 'r') as oz:
        for name in oz.namelist():
            output_files[name] = oz.read(name)

    # 1. Copy every VML drawing + its .rels and every image media file from
    #    the template that the output is missing.
    for name in template_files:
        if (
            name.startswith('xl/drawings/vmlDrawing')
            or name.startswith('xl/drawings/_rels/vmlDrawing')
            or name.startswith('xl/media/')
        ) and name not in output_files:
            output_files[name] = template_files[name]

    # 2. Map sheet names → file paths in both workbooks (needed so we patch
    #    the *same logical sheet* even if openpyxl reordered or renumbered
    #    sheetN.xml on save).
    template_sheets = _parse_workbook_sheets(template_files['xl/workbook.xml'])
    output_sheets = _parse_workbook_sheets(output_files['xl/workbook.xml'])

    template_wb_rels = {r['Id']: r['Target'] for r in _parse_rels(template_files['xl/_rels/workbook.xml.rels'])}
    output_wb_rels = {r['Id']: r['Target'] for r in _parse_rels(output_files['xl/_rels/workbook.xml.rels'])}

    template_sheet_path = {name: template_wb_rels[rid] for name, rid in template_sheets if rid in template_wb_rels}
    output_sheet_path = {name: output_wb_rels[rid] for name, rid in output_sheets if rid in output_wb_rels}

    def _norm(p: str) -> str:
        return p[1:] if p.startswith('/') else ('xl/' + p if not p.startswith('xl/') else p)

    # 3. For each shared sheet, find the template's legacyDrawingHF target
    #    (the VML drawing it pointed at) and re-link the output sheet to the
    #    same VML drawing path.
    for sheet_name, t_path in template_sheet_path.items():
        if sheet_name not in output_sheet_path:
            continue
        t_sheet = _norm(t_path)
        o_sheet = _norm(output_sheet_path[sheet_name])
        t_rels_path = t_sheet.replace('worksheets/', 'worksheets/_rels/') + '.rels'
        o_rels_path = o_sheet.replace('worksheets/', 'worksheets/_rels/') + '.rels'
        if t_rels_path not in template_files:
            continue

        # Find the VML drawing the template's sheet points at.
        legacy_target = None
        for r in _parse_rels(template_files[t_rels_path]):
            if r.get('Type', '').endswith('/vmlDrawing'):
                legacy_target = r['Target']
                break
        if legacy_target is None:
            continue

        # Pick a relationship id that doesn't collide with output's existing rels.
        existing = _parse_rels(output_files.get(o_rels_path, b'<?xml version="1.0"?><Relationships xmlns="' + _REL_NS.encode() + b'"/>'))
        used_ids = {r['Id'] for r in existing}
        new_rid = next(f'rId{i}' for i in range(100, 999) if f'rId{i}' not in used_ids)

        # Inject the relationship into the output sheet's .rels.
        rels_xml = output_files[o_rels_path].decode('utf-8') if o_rels_path in output_files else (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{_REL_NS}"></Relationships>'
        )
        new_rel = (
            f'<Relationship Id="{new_rid}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
            f'Target="{legacy_target}"/>'
        )
        rels_xml = rels_xml.replace('</Relationships>', new_rel + '</Relationships>')
        output_files[o_rels_path] = rels_xml.encode('utf-8')

        # Inject <legacyDrawingHF r:id="…"/> into the output sheet XML, just
        # before <tableParts> if present (legacyDrawingHF must precede
        # tableParts per the OOXML schema), else before </worksheet>.
        sheet_xml = output_files[o_sheet].decode('utf-8')
        if '<legacyDrawingHF' not in sheet_xml:
            tag = f'<legacyDrawingHF xmlns:r="{_R_NS}" r:id="{new_rid}"/>'
            if '<tableParts' in sheet_xml:
                sheet_xml = sheet_xml.replace('<tableParts', tag + '<tableParts', 1)
            else:
                sheet_xml = sheet_xml.replace('</worksheet>', tag + '</worksheet>')
            output_files[o_sheet] = sheet_xml.encode('utf-8')

    # 4. Make sure [Content_Types].xml registers the VML and PNG content
    #    types — openpyxl drops them when there's nothing in its model that
    #    needs them.
    ct_path = '[Content_Types].xml'
    if ct_path in output_files:
        ct = output_files[ct_path].decode('utf-8')
        added = []
        if 'ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"' not in ct:
            added.append('<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>')
        if 'Extension="png"' not in ct:
            added.append('<Default Extension="png" ContentType="image/png"/>')
        if added:
            ct = ct.replace('</Types>', ''.join(added) + '</Types>')
            output_files[ct_path] = ct.encode('utf-8')

    # 5. Re-zip.
    buf = io.BytesIO()
    with ZipFile(buf, 'w', ZIP_DEFLATED) as out_zip:
        for name, data in output_files.items():
            out_zip.writestr(name, data)
    return buf.getvalue()


# ───────────────────────────────────────────────────────────────────────
# Printer-settings / page-setup restoration (post-openpyxl-save)
# ───────────────────────────────────────────────────────────────────────

# The orientation Excel actually honours lives in the DEVMODE blob inside
# xl/printerSettings/printerSettingsN.bin, which each sheet's <pageSetup>
# references via r:id. openpyxl keeps the bare `orientation` attribute but drops
# the .bin part and the relationship, so a landscape template prints portrait.
# This restores the printerSettings parts and re-links each <pageSetup> to them
# — matching sheets by name (openpyxl may renumber sheetN.xml on save). It is a
# no-op for sheets/templates that carry no printerSettings.

_PRINTER_REL_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/printerSettings'


def _norm_xl_path(p: str) -> str:
    return p[1:] if p.startswith('/') else ('xl/' + p if not p.startswith('xl/') else p)


def restore_print_setup(template_bytes: bytes, output_bytes: bytes) -> bytes:
    template_files: dict[str, bytes] = {}
    with ZipFile(io.BytesIO(template_bytes), 'r') as tz:
        for name in tz.namelist():
            template_files[name] = tz.read(name)
    output_files: dict[str, bytes] = {}
    with ZipFile(io.BytesIO(output_bytes), 'r') as oz:
        for name in oz.namelist():
            output_files[name] = oz.read(name)

    # 1. Copy every printerSettings .bin the output is missing.
    for name in template_files:
        if name.startswith('xl/printerSettings/') and name not in output_files:
            output_files[name] = template_files[name]

    # 2. Map sheet name → sheet-part path in both workbooks (so we re-link the
    #    same logical sheet even if openpyxl renumbered sheetN.xml).
    template_sheets = _parse_workbook_sheets(template_files['xl/workbook.xml'])
    output_sheets = _parse_workbook_sheets(output_files['xl/workbook.xml'])
    template_wb_rels = {r['Id']: r['Target'] for r in _parse_rels(template_files['xl/_rels/workbook.xml.rels'])}
    output_wb_rels = {r['Id']: r['Target'] for r in _parse_rels(output_files['xl/_rels/workbook.xml.rels'])}
    template_sheet_path = {name: template_wb_rels[rid] for name, rid in template_sheets if rid in template_wb_rels}
    output_sheet_path = {name: output_wb_rels[rid] for name, rid in output_sheets if rid in output_wb_rels}

    for sheet_name, t_path in template_sheet_path.items():
        if sheet_name not in output_sheet_path:
            continue
        t_sheet = _norm_xl_path(t_path)
        o_sheet = _norm_xl_path(output_sheet_path[sheet_name])
        t_rels_path = t_sheet.replace('worksheets/', 'worksheets/_rels/') + '.rels'
        o_rels_path = o_sheet.replace('worksheets/', 'worksheets/_rels/') + '.rels'
        if t_rels_path not in template_files or o_sheet not in output_files:
            continue

        # The printerSettings part this sheet pointed at in the template.
        printer_target = next(
            (r['Target'] for r in _parse_rels(template_files[t_rels_path])
             if r.get('Type') == _PRINTER_REL_TYPE),
            None,
        )
        if printer_target is None:
            continue

        # Skip if openpyxl somehow preserved the link already.
        existing = _parse_rels(output_files.get(
            o_rels_path,
            b'<?xml version="1.0"?><Relationships xmlns="' + _REL_NS.encode() + b'"/>',
        ))
        if any(r.get('Type') == _PRINTER_REL_TYPE for r in existing):
            continue

        used_ids = {r['Id'] for r in existing}
        new_rid = next(f'rId{i}' for i in range(100, 999) if f'rId{i}' not in used_ids)

        # Inject the relationship into the output sheet's .rels.
        rels_xml = output_files[o_rels_path].decode('utf-8') if o_rels_path in output_files else (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{_REL_NS}"></Relationships>'
        )
        new_rel = f'<Relationship Id="{new_rid}" Type="{_PRINTER_REL_TYPE}" Target="{printer_target}"/>'
        rels_xml = rels_xml.replace('</Relationships>', new_rel + '</Relationships>')
        output_files[o_rels_path] = rels_xml.encode('utf-8')

        # Re-link <pageSetup> to it. openpyxl emits the tag with no r:id and no
        # xmlns:r on the root, so declare the prefix on the element itself.
        sheet_xml = output_files[o_sheet].decode('utf-8')
        m = re.search(r'<pageSetup\b[^>]*?>', sheet_xml)
        if m and 'r:id=' not in m.group(0):
            tag = m.group(0)
            close = '/>' if tag.endswith('/>') else '>'
            new_tag = f'{tag[:-len(close)]} r:id="{new_rid}" xmlns:r="{_R_NS}"{close}'
            output_files[o_sheet] = sheet_xml.replace(tag, new_tag, 1).encode('utf-8')

    # 3. Make sure [Content_Types].xml registers the .bin part type.
    ct_path = '[Content_Types].xml'
    if ct_path in output_files:
        ct = output_files[ct_path].decode('utf-8')
        if 'Extension="bin"' not in ct:
            ct = ct.replace(
                '</Types>',
                '<Default Extension="bin" '
                'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings"/>'
                '</Types>',
            )
            output_files[ct_path] = ct.encode('utf-8')

    buf = io.BytesIO()
    with ZipFile(buf, 'w', ZIP_DEFLATED) as out_zip:
        for name, data in output_files.items():
            out_zip.writestr(name, data)
    return buf.getvalue()
