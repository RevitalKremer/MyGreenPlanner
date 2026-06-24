"""Generate the Hebrew production-instructions workbook (.xlsx) for a project.

Two sheets, both filled from `app/templates/production_template.xlsx`:

  • "מסור" (Saw)     — saw-cut list for trapezoid members, external diagonals
                       and rails, split into three template sections.
  • "מנקבת" (Puncher) — punch operations for trapezoid frames only.

Like the price proposal, the template uses literal placeholder strings as the
contract between the file and this code (CUSTOMER_NAME / PROJECT_NAME /
PROPOSAL_DATE for the header, plus one anchor token per data section). The
shared fill machinery lives in `excel_template_utils`.

All cut/punch data is read straight from the project's persisted step-3
computed data (computedAreas + computedTrapezoids) — the same source the BOM
is built from — so the workshop list always matches the ordered material.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from app.services.bom_service import (
    _flatten_row_dict,
    _get_area_field,
    _group_pieces_by_length,
)
from app.services.excel_template_utils import (
    col_to_idx,
    extend_print_area,
    fill_section,
    find_anchor_row,
    replace_placeholders,
    restore_header_footer_images,
    restore_print_setup,
    set_table_range,
)


_TEMPLATES_DIR = Path(__file__).resolve().parents[1] / 'templates'
TEMPLATE_PATH = _TEMPLATES_DIR / 'production_template.xlsx'

# Header tokens shared with the proposal template (same names, same meaning).
_PLACEHOLDER_FIELDS = {'CUSTOMER_NAME', 'PROJECT_NAME', 'PROPOSAL_DATE'}

# Saw sheet: header row 10, then three label+anchor section pairs.
_SAW_SHEET = 'מסור'
_SAW_HEADER_ROW = 10
_SAW_TRAP_ANCHOR = 'SAW_TABLE_TRAP_DATA_ROW'
_SAW_EXT_DIAG_ANCHOR = 'SAW_TABLE_EXT_DIAG_DATA_ROW'
_SAW_RAILS_ANCHOR = 'SAW_TABLE_RAILS_DATA_ROW'
_SAW_COLUMN_MAP = {
    'approval': 'B',  # leftmost table column — filled with an unchecked ☐ per row
    'area':     'C',
    'trap':     'D',
    'desc':     'E',
    'length':   'F',
    'qty':      'G',
}

# Puncher sheet: header row 10, single section.
_PUNCH_SHEET = 'מנקבת'
_PUNCH_ANCHOR = 'PUNCH_TABLE_TRAP_DATA_ROW'
# Punch positions occupy ten dedicated columns (H..Q). Slots 1-9 each hold a
# single position; slot 10 holds the 10th position and every one beyond it,
# comma-joined (see _punch_slots).
_PUNCH_SLOT_COUNT = 10
_PUNCH_COLUMN_MAP = {
    'approval': 'B',  # leftmost table column — filled with an unchecked ☐ per row
    'area':     'C',
    'trap':     'D',
    'desc':     'E',
    'length':   'F',
    'qty':      'G',
    'punch1':   'H',
    'punch2':   'I',
    'punch3':   'J',
    'punch4':   'K',
    'punch5':   'L',
    'punch6':   'M',
    'punch7':   'N',
    'punch8':   'O',
    'punch9':   'P',
    'punch10':  'Q',
    'side':     'R',
    'diameter': 'S',
}

# Hebrew part labels (matched between the two sheets — the punch sheet's first
# five columns mirror the saw sheet's trapezoid rows).
_LABEL_BASE = 'בסיס'
_LABEL_SLOPE = 'יתר'
_LABEL_LEG = 'רגל'
_LABEL_DIAG = 'רגל אלכסונית'
_LABEL_DOUBLE = ' (כפול)'

# Punch type → (angle side, hole diameter). Keyed by the punch `origin`. Frame
# punches (legs / internal diagonals connecting at the beam) sit on side 1;
# block + rail punches sit on side 2. Diameter is a label, since leg/diagonal
# punches (built separately, below) are oval rather than round.
_PUNCH_TYPE_MAP = {
    'outerLeg':  {'side': 1, 'diameter': '9mm'},
    'innerLeg':  {'side': 1, 'diameter': '9mm'},
    'diagonal':  {'side': 1, 'diameter': '9mm'},
    'rail':      {'side': 2, 'diameter': '9mm'},
    'block':     {'side': 2, 'diameter': '13mm'},
    'connector': {'side': 1, 'diameter': '9mm'},  # splice bolt hole at a spliced-beam joint
}
_PUNCH_DEFAULT = {'side': 1, 'diameter': '9mm'}

# Leg & internal-diagonal punches are oval; they are not present in the
# `punches[]` array (which only records base/slope punches), so they are
# emitted directly from the legs[]/diagonals[] arrays with an empty position
# (filled in by the worker on the floor).
_OVAL = 'אובאלי'

# Unchecked checkbox placed in the אישור column of every data row, for the
# worker to tick by hand on the printed sheet (U+2610 BALLOT BOX).
_CHECKBOX = '☐'


# ───────────────────────────────────────────────────────────────────────
# Number / list formatting
# ───────────────────────────────────────────────────────────────────────

def _round_cm(v) -> float:
    """Round a length to 0.1 cm, returning an int when it lands whole."""
    r = round(float(v or 0), 1)
    return int(r) if r == int(r) else r


def _punch_slots(positions: list[float]) -> dict[str, object]:
    """Sorted, de-duplicated punch positions (cm) spread across the ten punch
    columns. Slots `punch1`..`punch9` each carry a single position; `punch10`
    carries the 10th position and every one beyond it, comma-joined. Returns a
    dict keyed by the slot names in `_PUNCH_COLUMN_MAP`; absent slots are left
    blank by the section filler."""
    seen: list[float] = []
    for p in sorted(_round_cm(p) for p in positions):
        if not seen or seen[-1] != p:
            seen.append(p)
    slots: dict[str, object] = {}
    for i in range(_PUNCH_SLOT_COUNT - 1):
        if i < len(seen):
            slots[f'punch{i + 1}'] = seen[i]
    overflow = seen[_PUNCH_SLOT_COUNT - 1:]
    if len(overflow) == 1:
        slots[f'punch{_PUNCH_SLOT_COUNT}'] = overflow[0]
    elif overflow:
        slots[f'punch{_PUNCH_SLOT_COUNT}'] = ', '.join(str(p) for p in overflow)
    return slots


# ───────────────────────────────────────────────────────────────────────
# Per-area / per-trapezoid extraction
# ───────────────────────────────────────────────────────────────────────

def _trap_instance_counts(computed_area: dict) -> dict[str, int]:
    """Count trapezoid instances in an area by counting its bases per
    trapezoidId (mirrors bom_service)."""
    counts: dict[str, int] = {}
    for b in _flatten_row_dict(computed_area.get('bases') or {}):
        tid = b.get('trapezoidId')
        if tid:
            counts[tid] = counts.get(tid, 0) + 1
    return counts


def _beam_parts(geom: dict, segments_key: str, kind: str, label: str,
                full_len: float, count: int) -> list[dict]:
    """Cut part(s) for one beam. A spliced beam (segments_key present with >1
    piece) yields one part per physical piece ('label 1/2', 'label 2/2', …),
    each tagged with `segmentIdx` so the punch builder can list that piece's own
    holes. An un-split beam yields a single part (segmentIdx 0)."""
    segs = geom.get(segments_key) or []
    if len(segs) > 1:
        n = len(segs)
        return [
            {'kind': kind, 'desc': f'{label} {i + 1}/{n}',
             'length': _round_cm(s.get('lengthCm', s['endCm'] - s['startCm'])),
             'qty': count, 'segmentIdx': i}
            for i, s in enumerate(segs)
        ]
    return [{'kind': kind, 'desc': label, 'length': full_len, 'qty': count, 'segmentIdx': 0}]


def _trap_parts(ct: dict, count: int) -> list[dict]:
    """Ordered list of cut parts for one trapezoid type, each carrying the five
    shared columns (area filled by the caller) plus a `kind` tag the punch
    builder uses. NOT grouped — one entry per beam piece / leg / diagonal."""
    geom = ct.get('geometry') or {}
    base_len = _round_cm(geom.get('baseBeamLength') or geom.get('baseLength') or 0)
    slope_len = _round_cm(geom.get('topBeamLength') or 0)

    parts: list[dict] = [
        *_beam_parts(geom, 'baseBeamSegments', 'base', _LABEL_BASE, base_len, count),
        *_beam_parts(geom, 'topBeamSegments', 'slope', _LABEL_SLOPE, slope_len, count),
    ]

    leg_n = 0
    for leg in ct.get('legs') or []:
        if leg.get('virtual'):
            continue
        leg_n += 1
        parts.append({
            'kind': 'leg',
            'desc': f'{_LABEL_LEG} {leg_n}',
            'length': _round_cm(leg.get('heightCm') or 0),
            'qty': count,
        })

    diag_n = 0
    for d in ct.get('diagonals') or []:
        if d.get('disabled'):
            continue
        diag_n += 1
        is_double = bool(d.get('isDouble'))
        parts.append({
            'kind': 'diagonal',
            'desc': f'{_LABEL_DIAG} {diag_n}' + (_LABEL_DOUBLE if is_double else ''),
            'length': _round_cm(d.get('lengthCm') or 0),
            'qty': count * (2 if is_double else 1),
        })

    return parts


def _iter_area_traps(data: dict):
    """Yield (area_label, computed_area, trapezoidId, computed_trapezoid, count)
    for every trapezoid instance group, ordered by area then trap id."""
    step2 = data.get('step2') or {}
    step3 = data.get('step3') or {}
    ca_by_label = {ca.get('label'): ca for ca in (step3.get('computedAreas') or [])}
    ct_by_id = {ct.get('trapezoidId'): ct for ct in (step3.get('computedTrapezoids') or [])}

    for idx, area in enumerate(step2.get('areas') or []):
        label = _get_area_field(area, 'label', f'Area {idx + 1}')
        ca = ca_by_label.get(label)
        if not ca:
            continue
        counts = _trap_instance_counts(ca)
        for tid in sorted(counts):
            count = counts[tid]
            ct = ct_by_id.get(tid)
            if not ct or count <= 0:
                continue
            yield label, ca, tid, ct, count


# ───────────────────────────────────────────────────────────────────────
# Row builders
# ───────────────────────────────────────────────────────────────────────

def build_saw_rows(data: dict) -> tuple[list[dict], list[dict], list[dict]]:
    """Return (trapezoid_rows, external_diagonal_rows, rail_rows) for the saw
    sheet. Material is kept per area — no cross-area consolidation."""
    trap_rows: list[dict] = []
    for label, _ca, tid, ct, count in _iter_area_traps(data):
        for part in _trap_parts(ct, count):
            trap_rows.append({
                'area': label,
                'trap': tid,
                'desc': part['desc'],
                'length': part['length'],
                'qty': part['qty'],
            })

    step2 = data.get('step2') or {}
    step3 = data.get('step3') or {}
    ca_by_label = {ca.get('label'): ca for ca in (step3.get('computedAreas') or [])}

    ext_rows: list[dict] = []
    rail_rows: list[dict] = []
    for idx, area in enumerate(step2.get('areas') or []):
        label = _get_area_field(area, 'label', f'Area {idx + 1}')
        ca = ca_by_label.get(label)
        if not ca:
            continue

        # External diagonals — one row per distinct cut length in this area.
        ext_pieces = [
            {'qty': 1, 'lenCm': (d.get('diagLengthMm') or 0) / 10}
            for d in (ca.get('diagonals') or [])
        ]
        for g in _group_pieces_by_length(ext_pieces, label, 'ext_diag'):
            ext_rows.append({
                'area': label, 'trap': '',
                'length': _round_cm(g['pieceLengthM'] * 100), 'qty': g['qty'],
            })

        # Rails — the raw stock cuts (mirrors bom_service rail_pieces: every
        # per-row rail's stockSegmentsMm, falling back to lengthCm).
        rail_pieces: list[dict] = []
        for r in _flatten_row_dict(ca.get('rails') or {}):
            segs = r.get('stockSegmentsMm') or []
            if segs:
                rail_pieces += [{'qty': 1, 'lenCm': s / 10} for s in segs if s > 0]
            elif (r.get('lengthCm') or 0) > 0:
                rail_pieces.append({'qty': 1, 'lenCm': r['lengthCm']})
        for g in _group_pieces_by_length(rail_pieces, label, 'rail'):
            rail_rows.append({
                'area': label, 'trap': '',
                'length': _round_cm(g['pieceLengthM'] * 100), 'qty': g['qty'],
            })

    for row in (*trap_rows, *ext_rows, *rail_rows):
        row['approval'] = _CHECKBOX
    return trap_rows, ext_rows, rail_rows


def build_punch_rows(data: dict) -> list[dict]:
    """Punch operations for trapezoid frames. Base & slope beams each emit two
    rows (one per angle side) from the punches[] array; legs and internal
    diagonals emit one oval-punch row each from their own arrays."""
    rows: list[dict] = []
    for label, _ca, tid, ct, count in _iter_area_traps(data):
        punches = ct.get('punches') or []
        for part in _trap_parts(ct, count):
            base5 = {
                'area': label, 'trap': tid,
                'desc': part['desc'], 'length': part['length'], 'qty': part['qty'],
            }
            if part['kind'] in ('base', 'slope'):
                beam_type = part['kind']  # punches[].beamType is 'base' | 'slope'
                seg_key = 'baseBeamSegments' if beam_type == 'base' else 'topBeamSegments'
                is_split = len((ct.get('geometry') or {}).get(seg_key) or []) > 1
                beam_punches = [p for p in punches if p.get('beamType') == beam_type]
                # On a split beam each part is one physical piece: restrict to that
                # piece's punches and measure each from the piece's own end.
                if is_split:
                    seg_idx = part.get('segmentIdx', 0)
                    beam_punches = [p for p in beam_punches if p.get('segmentIdx', 0) == seg_idx]
                    pos_key = 'piecePositionCm'
                else:
                    pos_key = 'positionCm'
                for side in (1, 2):
                    side_punches = [
                        p for p in beam_punches
                        if _PUNCH_TYPE_MAP.get(p.get('origin'), _PUNCH_DEFAULT)['side'] == side
                    ]
                    diameters = sorted({
                        _PUNCH_TYPE_MAP.get(p.get('origin'), _PUNCH_DEFAULT)['diameter']
                        for p in side_punches
                    })
                    rows.append({
                        **base5,
                        **_punch_slots([p.get(pos_key, p.get('positionCm', 0)) for p in side_punches]),
                        'side': side,
                        'diameter': ', '.join(diameters),
                    })
            else:  # leg / diagonal — single oval punch row, positions left blank
                rows.append({
                    **base5,
                    'side': '',
                    'diameter': _OVAL,
                })
    for row in rows:
        row['approval'] = _CHECKBOX
    return rows


# ───────────────────────────────────────────────────────────────────────
# Sheet fillers
# ───────────────────────────────────────────────────────────────────────

def _center_checkboxes(ws, col_letter: str, r0: int, r1: int) -> None:
    """Center the unchecked-checkbox cells in the אישור column so they read as
    a tick-box on the printed sheet."""
    col = col_to_idx(col_letter)
    for r in range(r0, r1 + 1):
        cell = ws.cell(r, col)
        if cell.value == _CHECKBOX:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def _fill_saw_sheet(ws, trap_rows, ext_rows, rail_rows, ctx) -> None:
    """Fill the saw sheet's three sections (which share one Excel table).

    Rows are inserted bottom-up so an upper section's insertion never
    invalidates a lower anchor's row number, then the single table + print
    area are grown once to cover the combined range.
    """
    replace_placeholders(ws, _PLACEHOLDER_FIELDS, ctx)
    sections = [
        (find_anchor_row(ws, _SAW_TRAP_ANCHOR), trap_rows),
        (find_anchor_row(ws, _SAW_EXT_DIAG_ANCHOR), ext_rows),
        (find_anchor_row(ws, _SAW_RAILS_ANCHOR), rail_rows),
    ]
    if any(anchor is None for anchor, _ in sections):
        raise ValueError(f"{ws.title}: a saw-sheet anchor placeholder was not found")

    # The approval column ships with a checkbox-prototype value (boolean False)
    # on every anchor row. A filled section overwrites it with ☐ on its first
    # data row, but an empty section would leave the stray False visible — clear
    # it up front so unused sections stay blank.
    approval_col = col_to_idx(_SAW_COLUMN_MAP['approval'])
    for anchor, _rows in sections:
        ws.cell(anchor, approval_col).value = None

    trap_anchor = sections[0][0]
    total_added = 0
    # Fill lowest anchor first so higher anchors keep their row numbers.
    for anchor, rows in sorted(sections, key=lambda s: s[0], reverse=True):
        if rows:
            fill_section(ws, anchor, rows, _SAW_COLUMN_MAP, extend_table=False)
            total_added += len(rows) - 1

    # The bottom-most data row = original table bottom (the rails anchor) plus
    # every row inserted above it; grow the table + print area to match.
    last_data_row = max(a for a, _ in sections) + total_added
    set_table_range(ws, _SAW_HEADER_ROW, last_data_row)
    if total_added:
        # Reuse extend_print_area: anchored above the footer, grow by total_added.
        extend_print_area(ws, trap_anchor, total_added + 1)
    _center_checkboxes(ws, _SAW_COLUMN_MAP['approval'], trap_anchor, last_data_row)


def _fill_punch_sheet(ws, rows, ctx) -> None:
    replace_placeholders(ws, _PLACEHOLDER_FIELDS, ctx)
    anchor = find_anchor_row(ws, _PUNCH_ANCHOR)
    if anchor is None:
        raise ValueError(f"{ws.title}: anchor placeholder {_PUNCH_ANCHOR!r} not found")
    # Clear the checkbox-prototype value (boolean False) so it never shows when
    # there are no rows; filled rows overwrite it with ☐.
    ws.cell(anchor, col_to_idx(_PUNCH_COLUMN_MAP['approval'])).value = None
    if not rows:
        return
    last_data_row = fill_section(ws, anchor, rows, _PUNCH_COLUMN_MAP)

    # Slots 1-9 each hold a single position, but the 10th column can carry an
    # overflow list (10th punch onward) when a beam side has more than ten
    # punches. The operator prints this sheet and punches each position, so wrap
    # that last column and drop any forced row height so Excel auto-fits it.
    loc_col = col_to_idx(_PUNCH_COLUMN_MAP[f'punch{_PUNCH_SLOT_COUNT}'])
    for r in range(anchor, last_data_row + 1):
        cell = ws.cell(r, loc_col)
        a = cell.alignment
        cell.alignment = Alignment(
            horizontal=a.horizontal, vertical='top', wrap_text=True,
            text_rotation=a.textRotation, indent=a.indent,
        )
        # Clear the copied prototype height so the row auto-fits the wrap.
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    _center_checkboxes(ws, _PUNCH_COLUMN_MAP['approval'], anchor, last_data_row)


# ───────────────────────────────────────────────────────────────────────
# Orchestrator
# ───────────────────────────────────────────────────────────────────────

def generate_production(project) -> bytes:
    """Build the production-instructions xlsx for `project` and return bytes.

    Reads the project's persisted step-3 computed data directly — no DB access,
    no product pricing — so it can run anywhere the project object is loaded.
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Production template not found at {TEMPLATE_PATH}")

    data = project.data or {}
    trap_rows, ext_rows, rail_rows = build_saw_rows(data)
    punch_rows = build_punch_rows(data)

    ctx = {
        'CUSTOMER_NAME': getattr(project, 'client_name', '') or '',
        'PROJECT_NAME':  project.name or '',
        'PROPOSAL_DATE': datetime.now(timezone.utc).strftime('%-d/%-m/%Y'),
    }

    wb = load_workbook(TEMPLATE_PATH)
    if _SAW_SHEET in wb.sheetnames:
        _fill_saw_sheet(wb[_SAW_SHEET], trap_rows, ext_rows, rail_rows, ctx)
    if _PUNCH_SHEET in wb.sheetnames:
        _fill_punch_sheet(wb[_PUNCH_SHEET], punch_rows, ctx)

    # The template ships with defined names that point at an external
    # `[1]ראשי!#REF!` cell plus the external-link entry; openpyxl re-emits them
    # in a shape Excel rejects, so strip them before serialising (same as the
    # proposal template).
    for name in list(wb.defined_names):
        if '#REF!' in (wb.defined_names[name].value or ''):
            del wb.defined_names[name]
    if getattr(wb, '_external_links', None):
        wb._external_links = []

    out = io.BytesIO()
    wb.save(out)
    template_bytes = TEMPLATE_PATH.read_bytes()
    result = restore_header_footer_images(template_bytes, out.getvalue())
    # openpyxl also drops the printerSettings part + <pageSetup r:id>, which
    # reverts the landscape template to portrait — restore them.
    return restore_print_setup(template_bytes, result)
