"""Generate a Hebrew price proposal (.xlsx) from a project's BOM.

Loads `app/templates/proposal_template.xlsx`, fills the two sheets
(`pricing` and `quantities`) with project metadata + BOM data, returns the
populated workbook as bytes for download.

The template uses literal placeholder strings inside cells as the contract
between the file and this code:
    CUSTOMER_NAME           – customer/client name
    PROJECT_NAME            – project name
    PROPOSAL_DATE           – proposal date (today)
    QTY_TABLE_DATA_ROW      – first BOM data row on the quantities sheet
    PRICE_TABLE_DATA_ROW    – first BOM data row on the pricing sheet

The placeholder cell is overwritten by the code; the row that holds it is
the styling prototype every BOM line is copied from. Rows below the data
anchor (totals, VAT, terms, signatures) are pushed down with insert_rows,
and SUM ranges in the totals row are rewritten to span the new data block.
"""
from __future__ import annotations

import asyncio
import io
import math
import os
import re
import shutil
import subprocess
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product
from app.models.user import User
from app.models.company import Company
from app.services import bom_service
from app.services.excel_template_utils import (
    col_to_idx,
    fill_section,
    find_anchor_row,
    replace_placeholders,
    restore_header_footer_images,
)


_TEMPLATES_DIR = Path(__file__).resolve().parents[1] / 'templates'
TEMPLATE_PATH = _TEMPLATES_DIR / 'proposal_template.xlsx'
_LOGO_PATH = _TEMPLATES_DIR / 'proposal_logo.png'
_FOOTER_PATH = _TEMPLATES_DIR / 'proposal_footer.png'


# Section labels in Hebrew (proposal is Hebrew-only regardless of app lang).
# We use the em-dash (U+2014) rather than the box-drawing horizontal (U+2500)
# the BOM Excel uses, because Noto Sans Hebrew (and most Hebrew fonts) don't
# include U+2500 — LibreOffice's PDF renderer falls back to a tofu rectangle.
_SECTION_LABEL_HE = {
    'rails':              '— קושרות —',
    'trapezoids':         '— טרפזים —',
    'diagonals_external': '— דיאגונלים —',
    'other':              '— אביזרי עזר —',
}

# Order rows are emitted in (mirrors the proposal template screenshot).
_SECTION_ORDER = ['rails', 'trapezoids', 'diagonals_external', 'other']


# ───────────────────────────────────────────────────────────────────────
# Row preparation
# ───────────────────────────────────────────────────────────────────────

def _classify(item: dict) -> str:
    section = item.get('section')
    if section in _SECTION_LABEL_HE:
        return section
    if item.get('element') == 'rail_40x40' and item.get('pieceLengthM') is not None:
        return 'rails'
    return 'other'


def _parse_extra_pct(raw) -> int:
    """Leading integer of the extra-percent value (mirrors the frontend
    parseExtraPct: handles None, numbers, and strings like '5' or '5%')."""
    if raw is None:
        return 0
    if isinstance(raw, (int, float)):
        return int(raw)
    m = re.match(r'\s*(-?\d+)', str(raw))
    return int(m.group(1)) if m else 0


def _effective_qty(item: dict) -> float:
    """Quantity to quote = base qty + extras (spares). Extras come from an
    explicit override when present, otherwise derived from extraPct — matching
    the frontend BOMView: extras = ceil(qty * pct / 100)."""
    qty = item.get('qty', 0) or 0
    extras = item.get('extras')
    if extras is None:
        extras = math.ceil(qty * _parse_extra_pct(item.get('extraPct')) / 100)
    return qty + extras


# Label shown in the notes column when a line was manually edited.
_EDITED_LABEL_HE = 'עודכן ידנית'


def _format_notes(item: dict) -> str:
    """Build the 'הערות' (notes) cell: the manual-edit marker and the free-text
    note, concatenated. Empty for untouched lines."""
    parts = []
    if item.get('edited'):
        parts.append(_EDITED_LABEL_HE)
    note = (item.get('note') or '').strip()
    if note:
        parts.append(note)
    return ' — '.join(parts)


def _build_data_row(line_num: int, item: dict, product: Product | None) -> dict:
    """Produce one filled-in row dict matching the template's column layout."""
    qty = _effective_qty(item)
    piece_length = item.get('pieceLengthM')
    is_length_row = piece_length is not None

    price_per_unit = (product.price_ils if product else None) or 0.0
    weight_per_unit = (product.weight_kg if product else None) or 0.0
    dep_pct = (product.depreciation_pct if product else None) or 0.0

    if is_length_row:
        # Length items: price and weight are per metre, multiplied by total length.
        total_length = piece_length * qty
        total_weight = total_length * weight_per_unit
        total_price = total_length * price_per_unit
        # Column G shows total weight for length items (kg).
        qty_kg_or_units = round(total_weight, 3)
    else:
        total_weight = qty * weight_per_unit
        total_price = qty * price_per_unit
        # Column G shows the piece count for piece items (units).
        qty_kg_or_units = qty

    # Depreciation: percent uplift on quantity → extra cost = total_price × dep%.
    dep_qty = round(qty * dep_pct / 100, 4) if dep_pct else 0
    total_with_dep = round(total_price * (1 + dep_pct / 100), 2) if dep_pct else round(total_price, 2)

    # Summary rows (depreciation_waste, processing) carry their amount in
    # `qty` (meters) with pieceLengthM=1 and zero weight. The per-row total
    # formula in column I is `unit_price × G (qty_kg_or_units)`, so the amount
    # must live in G — otherwise it multiplies by the zero weight and totals 0.
    # Shift the displayed values one cell left: blank the length column, put a
    # count of 1 in `qty` (F), and the meters in G.
    is_summary = item.get('element') in ('depreciation_waste', 'processing')

    return {
        'is_section_header':       False,
        'line_num':                line_num,
        'location':                item.get('areaLabel') or '',
        'product_name_he':         (product.name_he if product else None) or item.get('name') or item.get('element', ''),
        'length_m':                None if is_summary else (round(piece_length, 2) if is_length_row else None),
        'qty':                     1 if is_summary else qty,
        'qty_kg_or_units':         round(qty, 2) if is_summary else qty_kg_or_units,
        'unit_price_ils':          round(price_per_unit, 2),
        'total_price_ils':         round(total_price, 2),
        'depreciation_qty':        dep_qty,
        'total_with_depreciation': total_with_dep,
        'weight_kg':               round(total_weight, 3),
        'notes':                   _format_notes(item),
    }


def _build_section_header_row(section_key: str) -> dict:
    return {
        'is_section_header': True,
        'line_num':          '---',
        'location':          '',
        'product_name_he':   _SECTION_LABEL_HE[section_key],
        'length_m':          None,
        'qty':               '',
        'qty_kg_or_units':   '',
        'unit_price_ils':    '',
        'total_price_ils':   '',
        'depreciation_qty':  '',
        'total_with_depreciation': '',
        'weight_kg':         '',
        'notes':             '',
    }


def build_proposal_rows(bom_items: list[dict], products_by_type: dict[str, Product]) -> list[dict]:
    """Group BOM items into the four proposal sections, prepend a section
    header to each non-empty group, and number data rows continuously
    (section headers carry '---' in the line# column, like the BOM Excel)."""
    grouped: dict[str, list[dict]] = {k: [] for k in _SECTION_ORDER}
    for it in bom_items:
        grouped[_classify(it)].append(it)

    rows: list[dict] = []
    line_num = 0
    for sect in _SECTION_ORDER:
        items = grouped[sect]
        if not items:
            continue
        rows.append(_build_section_header_row(sect))
        for it in items:
            line_num += 1
            rows.append(_build_data_row(line_num, it, products_by_type.get(it.get('element'))))
    return rows


# ───────────────────────────────────────────────────────────────────────
# Template fill
# ───────────────────────────────────────────────────────────────────────

# Maps row dict keys to template column letters (RTL: B is leftmost in the
# physical XML, but renders rightmost in Excel — that's the template author's
# concern, not ours; we just write into the columns the template already has).
# The two sheets differ: the 'הערות' (notes) column was inserted at J on the
# pricing sheet (shifting depreciation/total/weight to K/L/M) but appended at M
# on the quantities sheet — so each sheet needs its own map.
_QTY_COLUMN_MAP = {
    'line_num':                'B',
    'location':                'C',
    'product_name_he':         'D',
    'length_m':                'E',
    'qty':                     'F',
    'qty_kg_or_units':         'G',
    'unit_price_ils':          'H',
    'total_price_ils':         'I',
    'depreciation_qty':        'J',
    'total_with_depreciation': 'K',
    'weight_kg':               'L',
    'notes':                   'M',
}

_PRICE_COLUMN_MAP = {
    'line_num':                'B',
    'location':                'C',
    'product_name_he':         'D',
    'length_m':                'E',
    'qty':                     'F',
    'qty_kg_or_units':         'G',
    'unit_price_ils':          'H',
    'total_price_ils':         'I',
    'notes':                   'J',
    'depreciation_qty':        'K',
    'total_with_depreciation': 'L',
    'weight_kg':               'M',
}

_PLACEHOLDER_FIELDS = {'CUSTOMER_NAME', 'PROJECT_NAME', 'PROPOSAL_DATE', 'ALUMINIUM_WEIGHT', 'BLOCKS_WEIGHT'}


def _apply_after_discount(ws, ctx: dict) -> None:
    """Replace the PRICE_AFTER_DISCOUNT token with a formula referencing the
    cell directly above it:
        `=<above>*(100-d)/100`   when a discount is set, or
        `=<above>`               when there is none (normal price).

    Must run AFTER row insertion/formula-shifting: insert_rows moves the token
    cell and the cell above it down together, so the relative `row-1`
    reference stays correct at the final position.
    """
    discount = ctx.get('DISCOUNT_PCT')
    for excel_row in ws.iter_rows():
        for cell in excel_row:
            if cell.value is None:
                continue
            if str(cell.value).strip() == 'PRICE_AFTER_DISCOUNT':
                col = get_column_letter(cell.column)
                above = cell.row - 1
                if discount is None or discount == 0:
                    cell.value = f'={col}{above}'
                else:
                    cell.value = f'={col}{above}*(100-{discount:g})/100'


def _autosize_note_rows(ws, anchor_row: int, rows: list[dict], column_map: dict) -> None:
    """Let rows with a note grow to fit it. The template pins data rows to a
    fixed (customHeight) 17pt, which clips the wrapped הערות text. For each row
    that actually has a note we enable wrap on the notes cell and release the
    fixed height so Excel / LibreOffice auto-fit the row to its content. Rows
    without a note keep the compact fixed height."""
    from openpyxl.styles import Alignment
    note_col = column_map.get('notes')
    if not note_col:
        return
    cidx = col_to_idx(note_col)
    for offset, row in enumerate(rows):
        if not (row.get('notes') or '').strip():
            continue
        excel_row = anchor_row + offset
        cell = ws.cell(excel_row, cidx)
        a = cell.alignment
        cell.alignment = Alignment(
            horizontal=a.horizontal, vertical=a.vertical, indent=a.indent,
            text_rotation=a.text_rotation, shrink_to_fit=a.shrink_to_fit,
            readingOrder=a.readingOrder, wrap_text=True,
        )
        # height=None drops the fixed customHeight so the row auto-fits content.
        ws.row_dimensions[excel_row].height = None


def _fill_sheet(ws, anchor_placeholder: str, rows: list[dict], ctx: dict, column_map: dict) -> None:
    # 1. Replace the static placeholders, then locate + clear the data anchor.
    replace_placeholders(ws, _PLACEHOLDER_FIELDS, ctx)
    anchor_row = find_anchor_row(ws, anchor_placeholder)
    if anchor_row is None:
        raise ValueError(f"{ws.title}: anchor placeholder {anchor_placeholder!r} not found")

    # 2. Insert + fill the BOM rows (no-op insertion when there are none, in
    #    which case the token stays at its template position).
    if rows:
        fill_section(ws, anchor_row, rows, column_map)
        _autosize_note_rows(ws, anchor_row, rows, column_map)

    # 3. Resolve the PRICE_AFTER_DISCOUNT token against the cell above it, now
    #    that every cell is at its final post-insertion position.
    _apply_after_discount(ws, ctx)


# ───────────────────────────────────────────────────────────────────────
# Orchestrator
# ───────────────────────────────────────────────────────────────────────

async def _load_products_by_type(db: AsyncSession) -> dict[str, Product]:
    result = await db.execute(select(Product).where(Product.active == True))
    return {p.type_key: p for p in result.scalars().all()}


async def generate_proposal(db: AsyncSession, project) -> bytes:
    """Build the price proposal xlsx for `project` and return the bytes."""
    bom = await bom_service.get_bom(db, project.id)
    if bom is None or bom_service.is_bom_stale(project.data or {}, bom):
        bom = await bom_service.compute_and_save_bom(db, project)

    deltas = ((project.data or {}).get('step5') or {}).get('bomDeltas') or {}
    effective_items = bom_service.apply_bom_deltas(bom.items, deltas)

    products_by_type = await _load_products_by_type(db)
    proposal_rows = build_proposal_rows(effective_items, products_by_type)

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Proposal template not found at {TEMPLATE_PATH}")
    wb = load_workbook(TEMPLATE_PATH)

    today = datetime.now(timezone.utc).strftime('%-d/%-m/%Y')

    aluminium_weight = 0.0
    blocks_weight = 0.0
    for item in effective_items:
        # Summary rows (depreciation_waste, processing) carry meters/units of
        # the aggregate, not real material — they must not be double-counted
        # into weight totals.
        if item.get('element') in ('depreciation_waste', 'processing'):
            continue
        product = products_by_type.get(item.get('element'))
        if not product or not product.weight_kg:
            continue
        qty = _effective_qty(item)
        piece_len = item.get('pieceLengthM')
        total_weight = (piece_len * qty * product.weight_kg) if piece_len is not None else (qty * product.weight_kg)
        ptype = product.product_type or ''
        element = item.get('element', '')
        if ptype == 'aluminium':
            aluminium_weight += total_weight
        elif element == 'block_50x24x15':
            blocks_weight += total_weight

    # Client discount comes from the project owner's COMPANY (admin-set).
    # None / 0 → normal price (no discount applied).
    owner = await db.get(User, project.owner_id)
    company = await db.get(Company, owner.company_id) if owner and owner.company_id else None
    discount = (
        float(company.discount_percent)
        if company is not None and company.discount_percent is not None
        else None
    )

    ctx = {
        'CUSTOMER_NAME':    getattr(project, 'client_name', '') or '',
        'PROJECT_NAME':     project.name or '',
        'PROPOSAL_DATE':    today,
        'ALUMINIUM_WEIGHT': round(aluminium_weight, 1),
        'BLOCKS_WEIGHT':    round(blocks_weight, 1),
        'DISCOUNT_PCT':     discount,
    }

    if 'quantities' in wb.sheetnames:
        _fill_sheet(wb['quantities'], 'QTY_TABLE_DATA_ROW', proposal_rows, ctx, _QTY_COLUMN_MAP)
    if 'pricing' in wb.sheetnames:
        _fill_sheet(wb['pricing'], 'PRICE_TABLE_DATA_ROW', proposal_rows, ctx, _PRICE_COLUMN_MAP)

    # The template was authored from another workbook and ships with five
    # defined names that point at an external `[1]ראשי!#REF!` cell, plus the
    # external-link entry itself. openpyxl re-emits these into the saved
    # workbook in a shape Excel rejects ("we found a problem with some
    # content"), so strip them out before serialising.
    for name in list(wb.defined_names):
        if '#REF!' in (wb.defined_names[name].value or ''):
            del wb.defined_names[name]
    if getattr(wb, '_external_links', None):
        wb._external_links = []

    out = io.BytesIO()
    wb.save(out)
    return restore_header_footer_images(TEMPLATE_PATH.read_bytes(), out.getvalue())


# ───────────────────────────────────────────────────────────────────────
# PDF generation (one PDF per sheet, via LibreOffice headless)
# ───────────────────────────────────────────────────────────────────────

# Sheet-name → suggested filename stub. Hebrew names rendered in the PDF
# come from the worksheet itself; this is just the file label.
_PDF_SHEETS = ('pricing', 'quantities')

# soffice cold-start timeout in seconds. Conversion of a small workbook
# typically completes in 3-5 s; we give it generous headroom.
_SOFFICE_TIMEOUT = 60


def _apply_pdf_table_stripes(ws) -> None:
    """LibreOffice's xlsx → PDF export doesn't render `TableStyleMedium15`'s
    row stripes visibly, so the printed PDF looks like a flat black-and-white
    grid even though the same xlsx in Excel shows clear zebra rows. Explicitly
    paint every other table data row with a light-teal fill so the PDF has
    visible striping regardless of how the renderer interprets table styles.
    """
    from openpyxl.styles import PatternFill
    from openpyxl.utils.cell import range_boundaries

    # Same colour TableStyleMedium15 uses (Excel theme accent1, ~30% tint).
    stripe = PatternFill(start_color='FFD9E5F1', end_color='FFD9E5F1', fill_type='solid')

    for table_name in list(ws.tables):
        table = ws.tables[table_name]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        except Exception:
            continue
        first_data = min_row + (table.headerRowCount or 1)
        for r in range(first_data, max_row + 1):
            if (r - first_data) % 2 == 0:  # 1st, 3rd, 5th data row → stripe
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(r, c)
                    # Only paint when the cell isn't already filled — skips
                    # the section-header rows etc. that picked up template
                    # styling we don't want to clobber.
                    if cell.fill is None or cell.fill.fill_type in (None, 'none'):
                        cell.fill = stripe


def _force_rtl_sheet(ws) -> None:
    """Lock the sheet into right-to-left mode for the PDF path.

    Two layers of fixes:

    1. Sheet-view RTL (`sheet_view.rightToLeft=True`) — column ordering.
    2. Cell-level reading order — explicitly stamp `readingOrder=2` (RTL)
       on every cell whose alignment leaves the order auto-detected. The
       template author only set explicit RTL on a handful of cells; the
       rest defaulted to `readingOrder=0`. LibreOffice's PDF renderer
       resolves `0` as LTR-context when a cell contains *any* neutral or
       Latin character (parens, colon, the ₪ symbol, digits inside a
       formula), placing those characters on the wrong side of the Hebrew
       text. Forcing `2` makes bidi resolve in an RTL paragraph context
       across the whole sheet, regardless of mixed content.
    """
    from openpyxl.styles import Alignment
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass
    for row in ws.iter_rows():
        for cell in row:
            a = cell.alignment
            if a.readingOrder == 2:
                continue
            cell.alignment = Alignment(
                horizontal=a.horizontal,
                vertical=a.vertical,
                indent=a.indent,
                wrap_text=a.wrap_text,
                shrink_to_fit=a.shrink_to_fit,
                text_rotation=a.text_rotation,
                relativeIndent=a.relativeIndent,
                justifyLastLine=a.justifyLastLine,
                readingOrder=2,
            )


def _force_pdf_page_layout(ws) -> None:
    """Force the whole table onto a single page width (margins respected),
    spilling vertically across as many pages as needed.

    The template ships with both `pageSetUpPr.fitToPage=true` and
    `pageSetup.scale=76`. Excel and LibreOffice resolve that conflict
    differently — Excel uses the explicit scale, LibreOffice uses the fit
    settings (which are unset, so it auto-fits). Either side of that picks
    a layout that doesn't match the other. We resolve it explicitly: clear
    the fixed scale and set fitToWidth=1 / fitToHeight=0 (unlimited),
    which both renderers honour identically — table fits to one page
    wide, height is whatever it needs to be.
    """
    if ws.sheet_properties.pageSetUpPr is not None:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.scale = None
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    # Template ships with ~1 mm left/right margins which leaves the table
    # almost flush with the page edge. Bump them so the printed table has
    # some breathing room (≈ 8 mm). Top/bottom unchanged so the header
    # block + footer image still sit where the template intends.
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3


def _embed_branding_images_for_pdf(ws) -> None:
    """LibreOffice's xlsx → PDF converter does not render Excel's
    legacyDrawingHF (print-time `&G` placeholder pointing to VML drawings),
    so the logo + footer don't show up in the PDF. As a workaround for the
    PDF path only, embed those same images as cell-anchored pictures on the
    sheet itself. The images live next to the template as standalone PNGs
    so they can be swapped without touching the xlsx. Sized to match the
    template's print-time VML dimensions (logo 5in×36pt; footer 805pt×116pt)
    so the PDF visually matches what Excel prints from the template.
    """
    from openpyxl.drawing.image import Image as XLImage

    if _LOGO_PATH.exists():
        # Logo: 5in × 36pt → 480 × 48 px @ 96 DPI. Anchored across the top
        # row block (rows 1-3 are whitespace placeholders in the template),
        # roughly centered on column D in the table's RTL layout.
        logo = XLImage(str(_LOGO_PATH))
        logo.width = 480
        logo.height = 48
        logo.anchor = 'D1'
        ws.add_image(logo)

    if _FOOTER_PATH.exists():
        # Footer spans the full printable width via TwoCellAnchor — the
        # image is bound to a cell rectangle covering the entire print
        # area (cols A through last) on a single row, so it always lines
        # up with the table edges regardless of fitToWidth scaling. Aspect
        # ratio is preserved by sizing the row height proportionally to
        # the print area's column-width sum.
        from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
        from PIL import Image as PILImage

        with PILImage.open(_FOOTER_PATH) as pil:
            aspect = pil.size[0] / pil.size[1]  # original w/h, e.g. ~6.97

        bounds = _parse_print_area_bounds(ws)
        min_col, _, max_col, max_row_pa = bounds if bounds else (1, 1, 9, ws.max_row)

        # Convert the print area's column-width-sum to points to set a row
        # height that keeps the image's aspect ratio. (Excel char width × 7
        # = pixels at 96 DPI; pixels × 0.75 = points.)
        col_chars = sum(
            (ws.column_dimensions.get(get_column_letter(c)).width
             if ws.column_dimensions.get(get_column_letter(c)) and
                ws.column_dimensions.get(get_column_letter(c)).width
             else 8.43)
            for c in range(min_col, max_col + 1)
        )
        col_pt = col_chars * 7 * 0.75
        row_h_pt = col_pt / aspect

        # Anchor the footer right after the print area's last content row
        # (template-defined, post-fill extension). Using ws.max_row would
        # leave a gap because some templates leave stray formatted/empty rows
        # past the print area, inflating max_row beyond actual content.
        footer_row = max_row_pa + 1
        ws.row_dimensions[footer_row].height = row_h_pt + 4  # small bottom pad

        footer = XLImage(str(_FOOTER_PATH))
        footer.anchor = TwoCellAnchor(
            _from=AnchorMarker(col=min_col - 1, colOff=0, row=footer_row - 1, rowOff=0),
            to=AnchorMarker(col=max_col, colOff=0, row=footer_row, rowOff=0),
            editAs='oneCell',
        )
        ws.add_image(footer)
        _ensure_print_area_includes_row(ws, footer_row)


def _parse_print_area_bounds(ws):
    """Return (min_col, min_row, max_col, max_row) for the worksheet's print
    area, or None if it isn't set."""
    pa = ws.print_area
    if not pa:
        return None
    pa_str = pa[0] if isinstance(pa, list) else pa
    if not pa_str:
        return None
    body = pa_str
    if '!' in body:
        _, body = body.rsplit('!', 1)
    from openpyxl.utils.cell import range_boundaries
    try:
        return range_boundaries(body.replace('$', ''))
    except Exception:
        return None


def _ensure_print_area_includes_row(ws, target_row: int) -> None:
    """Grow the print area's bottom row to at least `target_row`,
    keeping its column range untouched."""
    pa = ws.print_area
    if not pa:
        return
    pa_str = pa[0] if isinstance(pa, list) else pa
    if not pa_str:
        return
    sheet_prefix = ''
    body = pa_str
    if '!' in body:
        sheet_prefix, body = body.rsplit('!', 1)
        sheet_prefix += '!'
    from openpyxl.utils.cell import range_boundaries, get_column_letter
    try:
        min_col, min_row, max_col, max_row = range_boundaries(body.replace('$', ''))
    except Exception:
        return
    if max_row >= target_row:
        return
    ws.print_area = (
        f"{sheet_prefix}${get_column_letter(min_col)}${min_row}:"
        f"${get_column_letter(max_col)}${target_row}"
    )


def _isolate_xlsx_to_sheets(full_xlsx_bytes: bytes, keep_sheets: list[str]) -> bytes:
    """Return xlsx bytes containing only the requested sheets in the given order.

    Drops all other sheets, applies PDF-rendering fixes (RTL, stripes, page
    layout, branding images) to each kept sheet, and re-embeds the header/
    footer VML images so they survive both LibreOffice PDF export and direct
    Excel printing.
    """
    keep_set = set(keep_sheets)
    wb = load_workbook(io.BytesIO(full_xlsx_bytes))
    for name in [s for s in wb.sheetnames if s not in keep_set]:
        del wb[name]
    for sheet_name in keep_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.sheet_state = 'visible'
        _force_rtl_sheet(ws)
        _apply_pdf_table_stripes(ws)
        _force_pdf_page_layout(ws)
        _embed_branding_images_for_pdf(ws)
    first = next((s for s in keep_sheets if s in wb.sheetnames), None)
    if first:
        wb.active = wb.sheetnames.index(first)
    out = io.BytesIO()
    wb.save(out)
    return restore_header_footer_images(TEMPLATE_PATH.read_bytes(), out.getvalue())


def _xlsx_to_pdf_sync(xlsx_bytes: bytes) -> bytes:
    """Run `soffice --headless --convert-to pdf` against the given xlsx.

    LibreOffice locks `~/.config/libreoffice` per process, so concurrent
    calls collide. We give each invocation its own UserInstallation in a
    private temp directory.
    """
    soffice = shutil.which('soffice') or shutil.which('libreoffice')
    if not soffice:
        raise RuntimeError(
            "LibreOffice (soffice) not found in PATH — cannot generate PDF. "
            "Make sure libreoffice-calc is installed in the BE container."
        )

    with tempfile.TemporaryDirectory(prefix='proposal-pdf-') as workdir:
        in_path = Path(workdir) / 'in.xlsx'
        out_dir = Path(workdir) / 'out'
        profile_dir = Path(workdir) / 'profile'
        out_dir.mkdir()
        profile_dir.mkdir()
        in_path.write_bytes(xlsx_bytes)

        result = subprocess.run(
            [
                soffice,
                f'-env:UserInstallation=file://{profile_dir}',
                '--headless',
                '--norestore',
                '--nolockcheck',
                '--nologo',
                '--convert-to', 'pdf',
                '--outdir', str(out_dir),
                str(in_path),
            ],
            capture_output=True,
            timeout=_SOFFICE_TIMEOUT,
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"soffice xlsx→pdf failed (rc={result.returncode}): "
                f"{result.stderr.decode('utf-8', errors='replace')[-500:]}"
            )

        pdfs = list(out_dir.glob('*.pdf'))
        if not pdfs:
            raise RuntimeError(
                "soffice produced no PDF; stderr: "
                f"{result.stderr.decode('utf-8', errors='replace')[-500:]}"
            )
        return pdfs[0].read_bytes()


async def generate_proposal_pdf(db: AsyncSession, project, sheets: list[str]) -> bytes:
    """Generate the proposal xlsx, keep only the requested sheets, convert to PDF."""
    valid = set(_PDF_SHEETS)
    clean = [s for s in sheets if s in valid]
    if not clean:
        raise ValueError(f"No valid sheets requested; expected a subset of {_PDF_SHEETS}")
    full_xlsx = await generate_proposal(db, project)
    filtered_xlsx = _isolate_xlsx_to_sheets(full_xlsx, clean)
    # soffice is blocking; offload to a thread so the FastAPI event loop stays
    # responsive while the (3-5 s) cold-start runs.
    return await asyncio.to_thread(_xlsx_to_pdf_sync, filtered_xlsx)
