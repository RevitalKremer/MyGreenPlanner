"""Electrical (Sadot goods) BOM service — fully separate from bom_service.

Owns the project_electrical_bom table end to end: compute, staleness, deltas,
and materialization. Shares NO state or code path with the construction BOM —
the two stacks are deliberately independent (different table, different service,
parallel endpoints).

The electrical BOM is built from:
  * Step-6 equipment selection (data.step6.inverters — inverters/batteries/…)
  * Step-7 string plan (data.step7.strings) — drives string-derived items

NOTE: precise cable cross-section / protection / grounding sizing per the
Israeli electrical code is a data spec not yet defined (see TIER2 plan). v1
emits the priced Sadot equipment rows (the major-cost items) and leaves a
clearly-marked extension point for string-derived material rows.
"""
import hashlib
import json
import uuid
from datetime import datetime, timezone

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from app.models.product import Product, SADOT_EQUIPMENT_TYPES
from app.models.electrical_bom import ProjectElectricalBOM

# Bump to invalidate all cached electrical BOMs.
_ELECTRICAL_BOM_LOGIC_VERSION = 1


# ── Staleness ────────────────────────────────────────────────────────────────

def compute_input_hash(data: dict) -> str:
    """SHA-256 of the electrical inputs that determine the BOM."""
    step6 = data.get('step6', {}) or {}
    step7 = data.get('step7', {}) or {}
    relevant = {
        '_v': _ELECTRICAL_BOM_LOGIC_VERSION,
        'inverters': step6.get('inverters', []),
        'batteries': step6.get('batteries', []),
        'settings': step6.get('settings', {}),
        'strings': step7.get('strings', []),
    }
    raw = json.dumps(relevant, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode()).hexdigest()


async def get_electrical_bom(db: AsyncSession, project_id: uuid.UUID) -> ProjectElectricalBOM | None:
    result = await db.execute(
        select(ProjectElectricalBOM).where(ProjectElectricalBOM.project_id == project_id)
    )
    return result.scalar_one_or_none()


def is_electrical_bom_stale(project_data: dict, bom: ProjectElectricalBOM) -> bool:
    return bom.input_hash != compute_input_hash(project_data)


# ── Product loading ──────────────────────────────────────────────────────────

async def _load_sadot_products_by_type(db: AsyncSession) -> dict[str, dict]:
    """Load active Sadot equipment products keyed by type_key."""
    result = await db.execute(
        select(Product).where(
            Product.active == True,
            Product.product_type.in_(SADOT_EQUIPMENT_TYPES),
        )
    )
    return {
        p.type_key: {
            'id': p.id,
            'type_key': p.type_key,
            'product_type': p.product_type,
            'part_number': p.part_number,
            'name': p.name,
            'name_he': p.name_he,
            'price_ils': p.price_ils,
            'sadot_url': p.sadot_url,
        }
        for p in result.scalars().all()
    }


# ── BOM construction ─────────────────────────────────────────────────────────

def _enrich(item: dict, products_by_type: dict[str, dict]) -> dict:
    """Read-through product fields onto a BOM row (admin edits propagate
    without a cache invalidation; the stored row is not mutated)."""
    prod = products_by_type.get(item.get('element'))
    out = {**item}
    if prod:
        out['productId'] = str(prod['id'])
        out['partNumber'] = prod.get('part_number')
        out['name'] = prod.get('name')
        out['nameHe'] = prod.get('name_he')
        out['priceIls'] = prod.get('price_ils')
        out['sadotUrl'] = prod.get('sadot_url')
        if not out.get('section'):
            out['section'] = prod.get('product_type')
    return out


def _string_derived_rows(data: dict, products_by_type: dict[str, dict]) -> list[dict]:
    """Cables / protection / grounding derived from the string layout.

    Deferred: the per-string cable-run lengths and protection ratings depend on
    the Israeli-code sizing spec (not yet defined). Returns [] for now; this is
    the single insertion point for those rows once the spec lands."""
    return []


def build_electrical_bom(data: dict, products_by_type: dict[str, dict]) -> list[dict]:
    """Build the electrical BOM rows from the Step-6 equipment selection (and,
    later, the Step-7 string plan)."""
    items: list[dict] = []

    # Equipment rows — one per selected Sadot product (inverters + batteries),
    # summing qty.
    step6 = data.get('step6', {}) or {}
    picks = (step6.get('inverters') or []) + (step6.get('batteries') or [])
    qty_by_key: dict[str, int] = {}
    for pick in picks:
        key = pick.get('typeKey')
        if not key:
            continue
        qty_by_key[key] = qty_by_key.get(key, 0) + max(1, int(pick.get('qty') or 1))
    for key, qty in qty_by_key.items():
        items.append(_enrich({'areaLabel': 'System', 'element': key, 'qty': qty}, products_by_type))

    items.extend(_string_derived_rows(data, products_by_type))
    return items


# ── Deltas (self-contained) ──────────────────────────────────────────────────

def _delta_key(item: dict) -> str:
    return f"{item['areaLabel']}||{item['element']}"


def apply_deltas(items: list[dict], deltas: dict) -> list[dict]:
    """Merge electrical bomDeltas (overrides/additions/alternatives) — the
    electrical stack's own copy, independent of the construction BOM."""
    overrides = deltas.get('overrides') or {}
    additions = deltas.get('additions') or []
    alternatives = deltas.get('alternatives') or {}

    effective = []
    for item in items:
        key = _delta_key(item)
        ov = overrides.get(key)
        if ov and ov.get('removed'):
            continue
        entry = {**item}
        if ov and 'qty' in ov:
            entry['qty'] = ov['qty']
        alt = alternatives.get(entry['element'])
        if alt:
            entry['altElement'] = alt
        effective.append(entry)

    for add in additions:
        effective.append({
            'areaLabel': add.get('areaLabel', 'System'),
            'element': add.get('element', ''),
            'qty': add.get('qty', 0),
            'productId': add.get('productId'),
            'partNumber': add.get('partNumber'),
            'name': add.get('name'),
            'section': add.get('section'),
        })
    return effective


# ── DB-aware orchestrators ───────────────────────────────────────────────────

async def compute_and_save_electrical_bom(db: AsyncSession, project) -> ProjectElectricalBOM:
    """Compute the electrical BOM from step6/step7 and upsert. Entering the
    electrical BOM step wipes any pending step9 deltas (canonical regen)."""
    data = project.data or {}
    products_by_type = await _load_sadot_products_by_type(db)
    items = build_electrical_bom(data, products_by_type)
    input_hash = compute_input_hash(data)

    step9 = data.get('step9') or {}
    if step9.get('bomDeltas'):
        step9['bomDeltas'] = {}
        data['step9'] = step9
        project.data = data
        flag_modified(project, 'data')

    existing = await get_electrical_bom(db, project.id)
    if existing:
        existing.items = items
        existing.input_hash = input_hash
        existing.updated_at = datetime.now(timezone.utc)
        flag_modified(existing, 'items')
        await db.commit()
        await db.refresh(existing)
        return existing

    bom = ProjectElectricalBOM(project_id=project.id, items=items, input_hash=input_hash)
    db.add(bom)
    await db.commit()
    await db.refresh(bom)
    return bom


async def materialize_electrical_bom(db: AsyncSession, project) -> ProjectElectricalBOM | None:
    """Recalc: apply step9 deltas onto the cached items, persist, clear deltas."""
    bom = await get_electrical_bom(db, project.id)
    if not bom:
        return None
    data = project.data or {}
    step9 = data.get('step9') or {}
    deltas = step9.get('bomDeltas') or {}

    products_by_type = await _load_sadot_products_by_type(db)
    base = [_enrich(it, products_by_type) for it in (bom.items or [])]
    effective = apply_deltas(base, deltas)

    # Commit alt-swaps into the stored element.
    for item in effective:
        alt = item.pop('altElement', None)
        if alt:
            item['element'] = alt
            prod = products_by_type.get(alt)
            if prod:
                item['productId'] = str(prod['id'])
                item['partNumber'] = prod.get('part_number')
                item['name'] = prod.get('name')
                item['nameHe'] = prod.get('name_he')
                item['priceIls'] = prod.get('price_ils')
                item['sadotUrl'] = prod.get('sadot_url')

    bom.items = effective
    bom.updated_at = datetime.now(timezone.utc)
    flag_modified(bom, 'items')

    if step9.get('bomDeltas'):
        step9['bomDeltas'] = {}
        data['step9'] = step9
        project.data = data
        flag_modified(project, 'data')

    await db.commit()
    await db.refresh(bom)
    return bom


async def reenrich_items(db: AsyncSession, items: list[dict]) -> list[dict]:
    if not items:
        return []
    products_by_type = await _load_sadot_products_by_type(db)
    return [_enrich(it, products_by_type) for it in items]


async def effective_items(db: AsyncSession, project) -> list[dict]:
    """Electrical BOM rows with step9 deltas applied — read-only (no writes)."""
    data = project.data or {}
    products_by_type = await _load_sadot_products_by_type(db)
    base = build_electrical_bom(data, products_by_type)
    deltas = (data.get('step9') or {}).get('bomDeltas') or {}
    return apply_deltas(base, deltas)


async def generate_electrical_proposal(db: AsyncSession, project) -> bytes:
    """Equipment price-proposal xlsx, filling the Sadot Energy חשבשבת template
    (`sadot_proposal_template.xlsx`). Reuses proposal_service's row-builder +
    column map; the equipment template shares the same column contract but uses
    SADOT_-prefixed anchor tokens."""
    from io import BytesIO
    from pathlib import Path
    from types import SimpleNamespace
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    from app.models.user import User
    from app.models.company import Company
    from app.services import proposal_service
    from app.services.excel_template_utils import (
        fill_section, find_anchor_row, replace_placeholders, restore_header_footer_images,
    )

    template = Path(__file__).resolve().parents[1] / 'templates' / 'sadot_proposal_template.xlsx'
    if not template.exists():
        raise FileNotFoundError(f"Equipment proposal template not found at {template}")

    items = await effective_items(db, project)
    # One proposal row per electrical BOM item. _build_data_row reads
    # price_ils / weight_kg / depreciation_pct / name_he off a product object;
    # wrap the BOM row's own fields so equipment items (no ORM product) fit.
    rows = [
        proposal_service._build_data_row(
            i, it,
            SimpleNamespace(price_ils=it.get('priceIls'), weight_kg=0,
                            depreciation_pct=0, name_he=it.get('nameHe'), name=it.get('name')),
        )
        for i, it in enumerate(items, start=1)
    ]

    wb = load_workbook(template)
    today = datetime.now(timezone.utc).strftime('%-d/%-m/%Y')
    owner = await db.get(User, project.owner_id) if project.owner_id else None
    company = await db.get(Company, owner.company_id) if owner and owner.company_id else None
    discount = float(company.discount_percent) if company and company.discount_percent is not None else None
    ctx = {
        'CUSTOMER_NAME': getattr(project, 'client_name', '') or '',
        'PROJECT_NAME': project.name or '',
        'PROPOSAL_DATE': today,
    }

    for sheet, anchor in (('pricing', 'SADOT_PRICE_TABLE_DATA_ROW'), ('quantities', 'SADOT_QTY_TABLE_DATA_ROW')):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        replace_placeholders(ws, {'CUSTOMER_NAME', 'PROJECT_NAME', 'PROPOSAL_DATE'}, ctx)
        anchor_row = find_anchor_row(ws, anchor)
        if anchor_row is not None and rows:
            fill_section(ws, anchor_row, rows, proposal_service._COLUMN_MAP)
        # Resolve SADOT_PRICE_AFTER_DISCOUNT against the cell above it.
        for excel_row in ws.iter_rows():
            for cell in excel_row:
                if isinstance(cell.value, str) and cell.value.strip() == 'SADOT_PRICE_AFTER_DISCOUNT':
                    col = get_column_letter(cell.column)
                    above = cell.row - 1
                    cell.value = f'={col}{above}' if not discount else f'={col}{above}*(100-{discount:g})/100'

    for name in list(wb.defined_names):
        if '#REF!' in (wb.defined_names[name].value or ''):
            del wb.defined_names[name]
    if getattr(wb, '_external_links', None):
        wb._external_links = []

    out = BytesIO()
    wb.save(out)
    return restore_header_footer_images(template.read_bytes(), out.getvalue())
